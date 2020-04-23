import numpy as np
import openpyxl
import estpop

sheet = openpyxl.load_workbook('data.xlsx').worksheets[0]

pops = {}
for i in range(1, sheet.max_row):
    code = sheet.cell(i + 1, 3).value

    if not code in pops:
        pops[code] = []

    males, females = [], []
    for j in range(29, 50):
        males.append(sheet.cell(i + 1, j).value)
        females.append(sheet.cell(i + 1, j + 22).value)
    pops[code].append([males, females])

ratios = {}
for k, v in pops.items():
    change_ratios, baby_ratios, tail_ratios = [], [], []
    try:
        for i in range(len(v) - 5):
            change_ratio, baby_ratio, tail_ratio = estpop.ratios(
                v[i], v[i + 5])
            change_ratios.append(change_ratio)
            baby_ratios.append(baby_ratio)
            tail_ratios.append(tail_ratio)

        ratios[k] = {
            'change_ratio': np.mean(change_ratios, axis=0).tolist(),
            'baby_ratio': float(np.mean(baby_ratios)),
            'tail_ratio': float(np.mean(tail_ratios))
        }
    except:
        pass

f = open('result.csv', mode='w')
f.write('code,year\n')

for k, v in pops.items():
    if k in [411, 421, 521]:
        change_ratio = ratios[0]['change_ratio']
        baby_ratio = ratios[0]['baby_ratio']
        tail_ratio = ratios[0]['tail_ratio']
    else:
        change_ratio = ratios[k]['change_ratio']
        baby_ratio = ratios[k]['baby_ratio']
        tail_ratio = ratios[k]['tail_ratio']

    try:
        year = 2020
        estimates = v[5]

        for i in range(7):
            estimates = estpop.simulate(estimates, change_ratio, baby_ratio,
                                        tail_ratio)
            f.write('%s,%s,%s,%s\n' % (k, year + i * 5, ','.join(
                map(str, estimates[0])), ','.join(map(str, estimates[1]))))
    except:
        print(k)

f.close()
